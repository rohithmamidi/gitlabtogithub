#!/usr/bin/env python3
"""
GitLab → GitHub Metadata Migration Script + Excel Reporting
"""

import os
import argparse
import logging
import datetime
import time
import jwt
import requests
import gitlab
from github import Github, Auth
from github.GithubException import GithubException
from openpyxl import Workbook
from openpyxl.styles import Font


# =============================
# CONFIG
# =============================

GITLAB_URL = "https://gitlab.com"
GITHUB_API = "https://api.github.com"
CURRENT_DATETIME = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


# =============================
# UTIL FUNCTIONS
# =============================

def log_and_print(message, level="info"):
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    if level == "error":
        logging.error(message)
        print(f"[ERROR {timestamp}] {message}")
    elif level == "success":
        logging.info(message)
        print(f"[SUCCESS {timestamp}] {message}")
    else:
        logging.info(message)
        print(f"[INFO {timestamp}] {message}")


def load_repositories_from_file(path):
    repos = []
    with open(path, "r", encoding="utf-8-sig") as f:
        for line in f:
            repo = line.strip()
            if repo:
                repos.append(repo)
    return repos


# =============================
# GITHUB APP AUTH
# =============================

def generate_github_app_token(app_id, installation_id, private_key_path):
    with open(private_key_path, "r") as f:
        private_key = f.read()

    now = int(time.time())
    payload = {
        "iat": now - 60,
        "exp": now + (10 * 60),
        "iss": app_id
    }

    encoded_jwt = jwt.encode(payload, private_key, algorithm="RS256")

    headers = {
        "Authorization": f"Bearer {encoded_jwt}",
        "Accept": "application/vnd.github+json"
    }

    url = f"{GITHUB_API}/app/installations/{installation_id}/access_tokens"
    response = requests.post(url, headers=headers)

    if response.status_code != 201:
        raise Exception(f"GitHub App token error: {response.text}")

    return response.json()["token"]


# =============================
# METADATA MIGRATION
# =============================

def migrate_metadata(gl_project, gh_repo):

    summary = {}

    # LABELS
    try:
        existing = [l.name for l in gh_repo.get_labels()]
        for label in gl_project.labels.list(get_all=True):
            if label.name not in existing:
                gh_repo.create_label(
                    name=label.name,
                    color=label.color.replace("#", ""),
                    description=label.description or ""
                )
        summary["labels"] = "completed"
    except Exception as e:
        summary["labels"] = f"failed: {e}"

    # MILESTONES
    try:
        existing = [m.title for m in gh_repo.get_milestones(state="all")]
        for ms in gl_project.milestones.list(get_all=True):
            if ms.title not in existing:
                gh_repo.create_milestone(
                    title=ms.title,
                    description=ms.description or "",
                    state="open" if ms.state == "active" else "closed"
                )
        summary["milestones"] = "completed"
    except Exception as e:
        summary["milestones"] = f"failed: {e}"

    # ISSUES
    try:
        existing_titles = [i.title for i in gh_repo.get_issues(state="all")]
        for issue in gl_project.issues.list(get_all=True):
            if issue.title not in existing_titles:
                gh_issue = gh_repo.create_issue(
                    title=issue.title,
                    body=issue.description or "",
                    labels=issue.labels or []
                )
                if issue.state == "closed":
                    gh_issue.edit(state="closed")
        summary["issues"] = "completed"
    except Exception as e:
        summary["issues"] = f"failed: {e}"

    # MERGE REQUESTS → PR
    try:
        existing_pr_titles = [p.title for p in gh_repo.get_pulls(state="all")]
        for mr in gl_project.mergerequests.list(get_all=True):
            if mr.title in existing_pr_titles:
                continue
            try:
                pr = gh_repo.create_pull(
                    title=mr.title,
                    body=(mr.description or "") + f"\n\nMigrated from GitLab MR #{mr.iid}",
                    head=mr.source_branch,
                    base=mr.target_branch
                )
                if mr.state in ["closed", "merged"]:
                    pr.edit(state="closed")
            except Exception:
                continue
        summary["merge_requests"] = "completed"
    except Exception as e:
        summary["merge_requests"] = f"failed: {e}"

    return summary


# =============================
# MAIN
# =============================

def main():

    parser = argparse.ArgumentParser(description="GitLab → GitHub Metadata Migration Tool")

    parser.add_argument("--gitlab-token", required=True)
    parser.add_argument("--gitlab-project")
    parser.add_argument("--gitlab-project-file")
    parser.add_argument("--github-org", required=True)
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--generate-report", action="store_true")

    auth_group = parser.add_mutually_exclusive_group(required=True)
    auth_group.add_argument("--github-token")
    auth_group.add_argument("--use-app", action="store_true")

    parser.add_argument("--github-app-id")
    parser.add_argument("--github-installation-id")
    parser.add_argument("--github-private-key")

    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    logging.basicConfig(
        filename=os.path.join(args.output_dir, f"migration_{CURRENT_DATETIME}.log"),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        force=True
    )

    # AUTH
    if args.use_app:
        github_token = generate_github_app_token(
            args.github_app_id,
            args.github_installation_id,
            args.github_private_key
        )
        auth = Auth.Token(github_token)
        gh = Github(auth=auth)
    else:
        auth = Auth.Token(args.github_token)
        gh = Github(auth=auth)

    gl = gitlab.Gitlab(GITLAB_URL, private_token=args.gitlab_token)

    if args.gitlab_project_file:
        projects = load_repositories_from_file(args.gitlab_project_file)
    else:
        projects = [args.gitlab_project]

    # Excel Setup
    if args.generate_report:
        wb = Workbook()
        ws = wb.active
        ws.title = "GitLab Repo Analysis"

        headers = [
            "Parent Group", "Subgroup", "Full Group Path", "Repository Name",
            "Repository URL", "Status", "Is Empty", "Visibility",
            "Last Activity", "Size (MB)", "Wiki Enabled", "Pipelines Exist",
            "Branch Count", "LFS Files Count", "Artifacts Count", "Primary Language",
            "Open PRs", "Closed PRs",
            "Open Issues", "Closed Issues",
            "Open Discussions", "Closed Discussions"
        ]

        ws.append(headers)
        for col in ws[1]:
            col.font = Font(bold=True)

    for project_name in projects:

        log_and_print(f"Starting migration: {project_name}")

        try:
            gl_project = gl.projects.get(project_name)
        except Exception as e:
            log_and_print(f"GitLab fetch failed: {e}", "error")
            continue

        repo_name = project_name.split("/")[-1]

        # Corrected GitHub repo handling
        try:
            gh_repo = gh.get_repo(f"{args.github_org}/{repo_name}")
            log_and_print("GitHub repo already exists")
        except GithubException:
            log_and_print("GitHub repo not found. Creating...")
            user = gh.get_user()
            gh_repo = user.create_repo(repo_name, private=True)
            log_and_print("Repository created under personal account", "success")

        migrate_metadata(gl_project, gh_repo)

        # Excel Reporting
        if args.generate_report:
            gl_project.refresh()
            stats = gl_project.attributes.get("statistics", {})

            parts = gl_project.path_with_namespace.split("/")
            parent = parts[0] if len(parts) > 0 else ""
            subgroup = parts[1] if len(parts) > 2 else ""

            languages = gl_project.languages()
            primary_language = max(languages, key=languages.get) if languages else "N/A"

            open_prs = len(gl_project.mergerequests.list(state="opened", get_all=True))
            closed_prs = len(gl_project.mergerequests.list(state="closed", get_all=True))
            open_issues = len(gl_project.issues.list(state="opened", get_all=True))
            closed_issues = len(gl_project.issues.list(state="closed", get_all=True))

            open_discussions = 0
            closed_discussions = 0

            for issue in gl_project.issues.list(get_all=True):
                count = len(issue.discussions.list(get_all=True))
                if issue.state == "opened":
                    open_discussions += count
                else:
                    closed_discussions += count

            for mr in gl_project.mergerequests.list(get_all=True):
                count = len(mr.discussions.list(get_all=True))
                if mr.state == "opened":
                    open_discussions += count
                else:
                    closed_discussions += count

            ws.append([
                parent,
                subgroup,
                gl_project.path_with_namespace,
                gl_project.name,
                gl_project.web_url,
                "Archived" if gl_project.archived else "Active",
                gl_project.empty_repo,
                gl_project.visibility,
                gl_project.last_activity_at,
                round(stats.get("repository_size", 0) / (1024 * 1024), 2),
                gl_project.wiki_enabled,
                len(gl_project.pipelines.list(per_page=1, get_all=False)) > 0,
                len(gl_project.branches.list(get_all=True)),
                stats.get("lfs_objects_size", 0),
                stats.get("job_artifacts_size", 0),
                primary_language,
                open_prs,
                closed_prs,
                open_issues,
                closed_issues,
                open_discussions,
                closed_discussions
            ])

    if args.generate_report:
        report_path = os.path.join(
            args.output_dir,
            f"gitlab_repo_analysis_{CURRENT_DATETIME}.xlsx"
        )
        wb.save(report_path)
        log_and_print(f"Excel report generated: {report_path}", "success")

    log_and_print("All migrations completed.", "success")


if __name__ == "__main__":
    main()



# #!/usr/bin/env python3
# """
# GitLab → GitHub Migration Script + Excel Reporting

# Features:
# - Create GitHub repository
# - Migrate labels
# - Migrate milestones
# - Migrate issues
# - Migrate merge requests (as PRs)
# - Optional GitLab repo Excel analysis report
# - Supports GitHub PAT or GitHub App authentication
# """

# import os
# import argparse
# import logging
# import datetime
# import time
# import jwt
# import requests
# import gitlab
# from github import Github
# from github.GithubException import GithubException
# from openpyxl import Workbook
# from openpyxl.styles import Font


# # =============================
# # CONFIG
# # =============================

# GITLAB_URL = "https://gitlab.com"
# GITHUB_API = "https://api.github.com"
# CURRENT_DATETIME = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


# # =============================
# # UTIL FUNCTIONS
# # =============================

# def mask_token(token, visible_chars=4):
#     if not token:
#         return ""
#     return token[:visible_chars] + "*" * (len(token) - visible_chars)


# def log_and_print(message, level="info"):
#     timestamp = datetime.datetime.now().strftime("%H:%M:%S")
#     if level == "error":
#         logging.error(message)
#         print(f"[ERROR {timestamp}] {message}")
#     elif level == "success":
#         logging.info(message)
#         print(f"[SUCCESS {timestamp}] {message}")
#     else:
#         logging.info(message)
#         print(f"[INFO {timestamp}] {message}")


# def load_repositories_from_file(path):
#     repos = []
#     with open(path, "r", encoding="utf-8-sig") as f:
#         for line in f:
#             repo = line.strip()
#             if repo:
#                 repos.append(repo)
#     return repos


# # =============================
# # GITHUB APP AUTH
# # =============================

# def generate_github_app_token(app_id, installation_id, private_key_path):
#     with open(private_key_path, "r") as f:
#         private_key = f.read()

#     now = int(time.time())
#     payload = {
#         "iat": now - 60,
#         "exp": now + (10 * 60),
#         "iss": app_id
#     }

#     encoded_jwt = jwt.encode(payload, private_key, algorithm="RS256")

#     headers = {
#         "Authorization": f"Bearer {encoded_jwt}",
#         "Accept": "application/vnd.github+json"
#     }

#     url = f"{GITHUB_API}/app/installations/{installation_id}/access_tokens"
#     response = requests.post(url, headers=headers)

#     if response.status_code != 201:
#         raise Exception(f"GitHub App token error: {response.text}")

#     return response.json()["token"]


# # =============================
# # METADATA MIGRATION
# # =============================

# def migrate_metadata(gl_project, gh_repo):
#     summary = {}

#     try:
#         existing = [l.name for l in gh_repo.get_labels()]
#         for label in gl_project.labels.list(all=True):
#             if label.name not in existing:
#                 gh_repo.create_label(
#                     name=label.name,
#                     color=label.color.replace("#", ""),
#                     description=label.description or ""
#                 )
#         summary["labels"] = "completed"
#     except Exception as e:
#         summary["labels"] = f"failed: {e}"

#     try:
#         existing = [m.title for m in gh_repo.get_milestones(state="all")]
#         for ms in gl_project.milestones.list(all=True):
#             if ms.title not in existing:
#                 gh_repo.create_milestone(
#                     title=ms.title,
#                     description=ms.description or "",
#                     state="open" if ms.state == "active" else "closed"
#                 )
#         summary["milestones"] = "completed"
#     except Exception as e:
#         summary["milestones"] = f"failed: {e}"

#     try:
#         existing_titles = [i.title for i in gh_repo.get_issues(state="all")]
#         for issue in gl_project.issues.list(all=True):
#             if issue.title not in existing_titles:
#                 gh_issue = gh_repo.create_issue(
#                     title=issue.title,
#                     body=issue.description or "",
#                     labels=issue.labels or []
#                 )
#                 if issue.state == "closed":
#                     gh_issue.edit(state="closed")
#         summary["issues"] = "completed"
#     except Exception as e:
#         summary["issues"] = f"failed: {e}"

#     try:
#         existing_pr_titles = [p.title for p in gh_repo.get_pulls(state="all")]
#         for mr in gl_project.mergerequests.list(all=True):
#             if mr.title in existing_pr_titles:
#                 continue
#             try:
#                 pr = gh_repo.create_pull(
#                     title=mr.title,
#                     body=(mr.description or "") + f"\n\nMigrated from GitLab MR #{mr.iid}",
#                     head=mr.source_branch,
#                     base=mr.target_branch
#                 )
#                 if mr.state in ["closed", "merged"]:
#                     pr.edit(state="closed")
#             except Exception:
#                 continue
#         summary["merge_requests"] = "completed"
#     except Exception as e:
#         summary["merge_requests"] = f"failed: {e}"

#     return summary


# # =============================
# # MAIN
# # =============================

# def main():
#     parser = argparse.ArgumentParser(description="GitLab → GitHub Migration Tool")

#     parser.add_argument("--gitlab-token", required=True)
#     parser.add_argument("--gitlab-project")
#     parser.add_argument("--gitlab-project-file")
#     parser.add_argument("--github-org", required=True)
#     parser.add_argument("--output-dir", default="output")
#     parser.add_argument("--generate-report", action="store_true")

#     auth_group = parser.add_mutually_exclusive_group(required=True)
#     auth_group.add_argument("--github-token")
#     auth_group.add_argument("--use-app", action="store_true")

#     parser.add_argument("--github-app-id")
#     parser.add_argument("--github-installation-id")
#     parser.add_argument("--github-private-key")

#     args = parser.parse_args()

#     os.makedirs(args.output_dir, exist_ok=True)

#     logging.basicConfig(
#         filename=os.path.join(args.output_dir, f"migration_{CURRENT_DATETIME}.log"),
#         level=logging.INFO,
#         format="%(asctime)s - %(levelname)s - %(message)s"
#     )

#     # Authentication
#     if args.use_app:
#         github_token = generate_github_app_token(
#             args.github_app_id,
#             args.github_installation_id,
#             args.github_private_key
#         )
#     else:
#         github_token = args.github_token

#     gl = gitlab.Gitlab(GITLAB_URL, private_token=args.gitlab_token)
#     gh = Github(github_token)

#     if args.gitlab_project_file:
#         projects = load_repositories_from_file(args.gitlab_project_file)
#     else:
#         projects = [args.gitlab_project]

#     # Excel setup
#     if args.generate_report:
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "GitLab Repo Analysis"

#         headers = [
#             "Parent Group", "Subgroup", "Full Group Path", "Repository Name",
#             "Repository URL", "Status", "Is Empty", "Visibility",
#             "Last Activity", "Size (MB)", "Wiki Enabled", "Pipelines Exist",
#             "Branch Count", "LFS Files Count", "Artifacts Count", "Primary Language",
#             "Open PRs", "Closed PRs",
#             "Open Issues", "Closed Issues",
#             "Open Discussions", "Closed Discussions"
#         ]

#         ws.append(headers)
#         for col in ws[1]:
#             col.font = Font(bold=True)

#     for project_name in projects:
#         log_and_print(f"Starting migration: {project_name}")

#         try:
#             gl_project = gl.projects.get(project_name)
#         except Exception as e:
#             log_and_print(f"GitLab fetch failed: {e}", "error")
#             continue

#         repo_name = project_name.split("/")[-1]
        
#         try:

#             gh_repo = gh.get_repo(f"{args.github_org}/{repo_name}")
#             log_and_print("GitHub repo already exists")
#         except GithubException:
#             log_and_print("GitHub repo not found. Creating...", "info")

#         try:
#              # Try as organization
#              org = gh.get_organization(args.github_org)
#              gh_repo = org.create_repo(repo_name, private=True)
#              log_and_print("Repository created under organization", "success")

#         except GithubException:
#             # Fallback to personal account
#             user = gh.get_user()
#             gh_repo = user.create_repo(repo_name, private=True)
#             log_and_print("Repository created under personal account", "success")


#         #     gh_repo = gh.get_repo(f"{args.github_org}/{repo_name}")
#         # except GithubException:
#         #     org = gh.get_organization(args.github_org)
#         #     gh_repo = org.create_repo(repo_name, private=True)

#         migrate_metadata(gl_project, gh_repo)

#         # Excel Data
#         if args.generate_report:
#             gl_project.refresh()
#             stats = gl_project.attributes.get("statistics", {})

#             parts = gl_project.path_with_namespace.split("/")
#             parent = parts[0] if len(parts) > 0 else ""
#             subgroup = parts[1] if len(parts) > 2 else ""

#             languages = gl_project.languages()
#             primary_language = max(languages, key=languages.get) if languages else "N/A"

#             open_prs = len(gl_project.mergerequests.list(state="opened", get_all=True))
#             closed_prs = len(gl_project.mergerequests.list(state="closed", get_all=True))

#             open_issues = len(gl_project.issues.list(state="opened", get_all=True))
#             closed_issues = len(gl_project.issues.list(state="closed", get_all=True))

#             # Discussions count
#             open_discussions = 0
#             closed_discussions = 0

#             for issue in gl_project.issues.list(state="opened", get_all=True):
#                 open_discussions += len(issue.discussions.list(get_all=True))

#             for issue in gl_project.issues.list(state="closed", get_all=True):
#                 closed_discussions += len(issue.discussions.list(get_all=True))

#             for mr in gl_project.mergerequests.list(state="opened", get_all=True):
#                 open_discussions += len(mr.discussions.list(get_all=True))

#             for mr in gl_project.mergerequests.list(state="closed", get_all=True):
#                 closed_discussions += len(mr.discussions.list(get_all=True))

#             ws.append([
#                 parent,
#                 subgroup,
#                 gl_project.path_with_namespace,
#                 gl_project.name,
#                 gl_project.web_url,
#                 "Archived" if gl_project.archived else "Active",
#                 gl_project.empty_repo,
#                 gl_project.visibility,
#                 gl_project.last_activity_at,
#                 round(stats.get("repository_size", 0) / (1024 * 1024), 2),
#                 gl_project.wiki_enabled,
#                 len(gl_project.pipelines.list(per_page=1, get_all=False)) > 0,
#                 len(gl_project.branches.list(get_all=True)),
#                 stats.get("lfs_objects_size", 0),
#                 stats.get("job_artifacts_size", 0),
#                 primary_language,
#                 open_prs,
#                 closed_prs,
#                 open_issues,
#                 closed_issues,
#                 open_discussions,
#                 closed_discussions
#             ])

#     if args.generate_report:
#         report_path = os.path.join(
#             args.output_dir,
#             f"gitlab_repo_analysis_{CURRENT_DATETIME}.xlsx"
#         )
#         wb.save(report_path)
#         log_and_print(f"Excel report generated: {report_path}", "success")

#     log_and_print("All migrations completed.", "success")


# if __name__ == "__main__":
#     main()


# #!/usr/bin/env python3
# """
# GitLab → GitHub Migration Script
# Author: Rohith (Improved & Corrected Version)

# Features:
# - Create GitHub repository
# - Migrate labels
# - Migrate milestones
# - Migrate issues
# - Migrate merge requests (as PRs)
# - Apply branch protection
# - Migrate webhooks
# - Migrate CI/CD variables → GitHub secrets
# - Supports GitHub PAT or GitHub App authentication
# """

# import os
# import argparse
# import logging
# import datetime
# import time
# import csv
# import jwt
# import requests
# import gitlab
# from github import Github
# from github.GithubException import GithubException


# # =============================
# # CONFIG
# # =============================

# GITLAB_URL = "https://gitlab.com"
# GITHUB_API = "https://api.github.com"
# CURRENT_DATETIME = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


# # =============================
# # UTIL FUNCTIONS
# # =============================

# def mask_token(token, visible_chars=4):
#     if not token:
#         return ""
#     return token[:visible_chars] + "*" * (len(token) - visible_chars)


# def log_and_print(message, level="info"):
#     timestamp = datetime.datetime.now().strftime("%H:%M:%S")
#     if level == "error":
#         logging.error(message)
#         print(f"[ERROR {timestamp}] {message}")
#     elif level == "success":
#         logging.info(message)
#         print(f"[SUCCESS {timestamp}] {message}")
#     else:
#         logging.info(message)
#         print(f"[INFO {timestamp}] {message}")


# def load_repositories_from_file(path):
#     repos = []
#     with open(path, "r", encoding="utf-8-sig") as f:
#         for line in f:
#             repo = line.strip()
#             if repo:
#                 repos.append(repo)
#     return repos


# # =============================
# # GITHUB APP AUTH
# # =============================

# def generate_github_app_token(app_id, installation_id, private_key_path):
#     with open(private_key_path, "r") as f:
#         private_key = f.read()

#     now = int(time.time())
#     payload = {
#         "iat": now - 60,
#         "exp": now + (10 * 60),
#         "iss": app_id
#     }

#     encoded_jwt = jwt.encode(payload, private_key, algorithm="RS256")

#     headers = {
#         "Authorization": f"Bearer {encoded_jwt}",
#         "Accept": "application/vnd.github+json"
#     }

#     url = f"{GITHUB_API}/app/installations/{installation_id}/access_tokens"
#     response = requests.post(url, headers=headers)

#     if response.status_code != 201:
#         raise Exception(f"GitHub App token error: {response.text}")

#     return response.json()["token"]


# # =============================
# # METADATA MIGRATION
# # =============================

# def migrate_metadata(gl_project, gh_repo):
#     summary = {}

#     # -------------------------
#     # LABELS
#     # -------------------------
#     try:
#         log_and_print("Migrating labels...")
#         existing = [l.name for l in gh_repo.get_labels()]

#         for label in gl_project.labels.list(all=True):
#             if label.name not in existing:
#                 gh_repo.create_label(
#                     name=label.name,
#                     color=label.color.replace("#", ""),
#                     description=label.description or ""
#                 )
#         summary["labels"] = "completed"
#     except Exception as e:
#         summary["labels"] = f"failed: {e}"

#     # -------------------------
#     # MILESTONES
#     # -------------------------
#     try:
#         log_and_print("Migrating milestones...")
#         existing = [m.title for m in gh_repo.get_milestones(state="all")]

#         for ms in gl_project.milestones.list(all=True):
#             if ms.title not in existing:
#                 gh_repo.create_milestone(
#                     title=ms.title,
#                     description=ms.description or "",
#                     state="open" if ms.state == "active" else "closed"
#                 )
#         summary["milestones"] = "completed"
#     except Exception as e:
#         summary["milestones"] = f"failed: {e}"

#     # -------------------------
#     # ISSUES
#     # -------------------------
#     try:
#         log_and_print("Migrating issues...")
#         existing_titles = [i.title for i in gh_repo.get_issues(state="all")]

#         for issue in gl_project.issues.list(all=True):
#             if issue.title not in existing_titles:
#                 gh_issue = gh_repo.create_issue(
#                     title=issue.title,
#                     body=issue.description or "",
#                     labels=issue.labels or []
#                 )
#                 if issue.state == "closed":
#                     gh_issue.edit(state="closed")
#         summary["issues"] = "completed"
#     except Exception as e:
#         summary["issues"] = f"failed: {e}"

#     # -------------------------
#     # MERGE REQUESTS → PRs
#     # -------------------------
#     try:
#         log_and_print("Migrating merge requests...")
#         existing_pr_titles = [p.title for p in gh_repo.get_pulls(state="all")]

#         for mr in gl_project.mergerequests.list(all=True):
#             if mr.title in existing_pr_titles:
#                 continue

#             try:
#                 pr = gh_repo.create_pull(
#                     title=mr.title,
#                     body=(mr.description or "") + f"\n\nMigrated from GitLab MR #{mr.iid}",
#                     head=mr.source_branch,
#                     base=mr.target_branch
#                 )
#                 if mr.state in ["closed", "merged"]:
#                     pr.edit(state="closed")
#             except Exception:
#                 continue

#         summary["merge_requests"] = "completed"
#     except Exception as e:
#         summary["merge_requests"] = f"failed: {e}"

#     return summary


# # =============================
# # MAIN
# # =============================

# def main():
#     parser = argparse.ArgumentParser(description="GitLab → GitHub Migration Tool")

#     parser.add_argument("--gitlab-token", required=True)
#     parser.add_argument("--gitlab-project")
#     parser.add_argument("--gitlab-project-file")
#     parser.add_argument("--github-org", required=True)
#     parser.add_argument("--output-dir", default="output")

#     auth_group = parser.add_mutually_exclusive_group(required=True)
#     auth_group.add_argument("--github-token")
#     auth_group.add_argument("--use-app", action="store_true")

#     parser.add_argument("--github-app-id")
#     parser.add_argument("--github-installation-id")
#     parser.add_argument("--github-private-key")

#     args = parser.parse_args()

#     os.makedirs(args.output_dir, exist_ok=True)

#     logging.basicConfig(
#         filename=os.path.join(args.output_dir, f"migration_{CURRENT_DATETIME}.log"),
#         level=logging.INFO,
#         format="%(asctime)s - %(levelname)s - %(message)s"
#     )

#     # =============================
#     # AUTH
#     # =============================

#     if args.use_app:
#         if not (args.github_app_id and args.github_installation_id and args.github_private_key):
#             raise ValueError("GitHub App parameters required")

#         github_token = generate_github_app_token(
#             args.github_app_id,
#             args.github_installation_id,
#             args.github_private_key
#         )
#         log_and_print(f"Using GitHub App token: {mask_token(github_token)}")
#     else:
#         github_token = args.github_token

#     gl = gitlab.Gitlab(GITLAB_URL, private_token=args.gitlab_token)
#     gh = Github(github_token)

#     # =============================
#     # PROJECT LIST
#     # =============================

#     if args.gitlab_project_file:
#         projects = load_repositories_from_file(args.gitlab_project_file)
#     else:
#         projects = [args.gitlab_project]

#     # =============================
#     # PROCESS EACH PROJECT
#     # =============================

#     for project_name in projects:
#         log_and_print(f"Starting migration: {project_name}")

#         try:
#             gl_project = gl.projects.get(project_name)
#         except Exception as e:
#             log_and_print(f"GitLab fetch failed: {e}", "error")
#             continue

#         repo_name = project_name.split("/")[-1]

#         try:
#             gh_repo = gh.get_repo(f"{args.github_org}/{repo_name}")
#             log_and_print("GitHub repo already exists")
#         except GithubException:
#             org = gh.get_organization(args.github_org)
#             gh_repo = org.create_repo(repo_name, private=True)
#             log_and_print("GitHub repo created", "success")

#         summary = migrate_metadata(gl_project, gh_repo)

#         log_and_print(f"Migration completed for {project_name}", "success")
#         log_and_print(f"Summary: {summary}")

#     log_and_print("All migrations completed.", "success")


# if __name__ == "__main__":
#     main()